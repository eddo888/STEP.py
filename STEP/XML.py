#!/usr/bin/env python3

# PYTHON_ARGCOMPLETE_OK

import sys, os, argparse, json, xmltodict, codecs, sqlalchemy

from datetime import datetime, date
from decimal import Decimal
from io import StringIO
from collections import namedtuple, OrderedDict
from openpyxl import load_workbook

from GoldenChild.xpath import *
from Perdy.pyxbext import directory
from Perdy.parser import printXML
from Spanners.IdentityCache import IdentityCache
from Argumental.Argue import Argue

from STEP.XSD import *

args = Argue()

#_________________________________________________________________
@args.command(single=True)
class Converter(object):
	'''
	tool to convert excel into step xml formats

	'''

	@args.property(default='PIM', help='prefix to use on ID definitions')
	def prefix(self): return

	@args.property(default='Context1', help='step context id')
	def context(self): return

	@args.property(default='Main', help='step workspace id')
	def workspace(self): return

	parent_type_id = 'Product user-type root'
	parent_product_id = 'Product hierarchy root'

	#_____________________________________________________________
	def __init__(self):
		self.ids = IdentityCache()

		self.types   = OrderedDict()
		self.ag      = None    # parent attribute group
		self.ags     = OrderedDict()
		self.lovs	 = OrderedDict()
		self.attrs   = OrderedDict()
		self.prods   = OrderedDict()

		self.doc = STEP_ProductInformation(
			ExportTime = datetime.now(),
			ContextID = self.context,
			WorkspaceID = self.workspace,
			UseContextLocale=False,
			ListOfValuesGroupList = ListOfValuesGroupListType(),
			ListsOfValues = ListsOfValuesType(),
			AttributeGroupList = AttributeGroupListType(),
			AttributeList = AttributeListType(),
			UserTypes = UserTypesType(),
			CrossReferenceTypes = CrossReferenceTypesType(),
			Products = ProductsType(),
			Classifications = ClassificationsType(),
		)


	#_____________________________________________________________
	def __del__(self):
		self.close()


	#_____________________________________________________________
	def close(self):
		if hasattr(self, 'ids'):
			self.ids.save()


	#_____________________________________________________________
	def save(self, file):
		with codecs.open(file, 'w', encoding='utf8') as output:
			printXML(self.doc.toxml(), output=output)


	#_____________________________________________________________
	def make_prod(self, id, name, type_id, parent_id):

		self.prods[id] = ProductType(
			ID = id,
			Name = [NameType(name)],
			UserTypeID = type_id,
			ParentID = parent_id,
		)
		self.doc.Products.append(self.prods[id])

		return self.prods[id]


	#_____________________________________________________________
	def make_type(self, id, name, parent_type_id):

		self.types[id] = UserTypeType(
			ID = id,
			Name = [NameType(name)],
			AllowInDesignTemplate='false',
			AllowQuarkTemplate='false',
			IDPattern='%s_[uuid]'%self.prefix,
			IsCategory='true',
			ManuallySorted='false',
			ReferenceTargetLockPolicy='Strict',
			Referenced='true',
			UserTypeLink = [
				UserTypeLinkType(
					UserTypeID=parent_type_id
				)
			]
		)
		self.doc.UserTypes.append(self.types[id])

		self.ags[id] = AttributeGroupType(
			ID = id,
			ShowInWorkbench = 'true',
			ManuallySorted = 'false',
			Name = [
				NameType(name)
			]
		)
		self.ag.AttributeGroup.append(self.ags[id])

		return self.types[id]


	#_____________________________________________________________
	def make_attr(self, id, name, attr_type, length, parent_group_id, user_type_id):

		if id not in self.attrs.keys():
			attr = AttributeType(
				ID = id,
				Name = [NameType(name)],
				MultiValued = 'false',
				ProductMode = 'Normal', #specification, use "Property" for description
				FullTextIndexed = 'false',
				ExternallyMaintained = 'false',
				Derived = 'false',

				AttributeGroupLink = [
					AttributeGroupLinkType(
						AttributeGroupID = parent_group_id
					)
				],

				MetaData = MetaDataType(
				),

				UserTypeLink = [
					UserTypeLinkType(
						UserTypeID = user_type_id
					)
				]
			)

			lov = None
			if lov:
				attr.ListOfValueLink = ListOfValueLinkType(
					ListOfValueID=lov
				)
			else:
				attr.Validation = ValidationType(
					BaseType = attr_type,
					MaxLength = length,
				)

			tooltip = None
			if tooltip:
				attr.MetaData.append(ValueType(
					tooltip,
					AttributeID='ToolTip'
				))

			self.doc.AttributeList.append(attr)
			self.attrs[id] = attr

		return self.attrs[id]


	#_____________________________________________________________
	def make_ref(self, id, name, source, target):

		reference = ProductCrossReferenceTypeType(
			ID = id,
			Name = [NameType(name)],
			UserTypeLink = [
				UserTypeLinkType(
					UserTypeID = source
				)
			],
			TargetUserTypeLink = [
				TargetUserTypeLinkType(
					UserTypeID = target
				)
			]
		)

		self.doc.CrossReferenceTypes.append(reference)


	#_____________________________________________________________
	def make_lov(self, id, name, tipe, length, parent='List Of Values group root'):
		if id not in self.lovs.keys():
			lov = ListOfValueType(
				ID = id,
				Name = [NameType(name)],
				ParentID = parent,
				UseValueID = 'true',
				AllowUserValueAddition = 'false',
				Validation = ValidationType(
					BaseType=tipe,
					MaxLength=length
				),
				Value = [
				]
			)
			self.doc.ListsOfValues.append(lov)
			self.lovs[id] = lov

		return self.lovs[id]


	#_____________________________________________________________
	@args.operation
	@args.parameter(name='id', help='LOV ID')
	@args.parameter(name='name', help='LOV Name')
	@args.parameter(name='input', help='excel file with attribute definitions')
	@args.parameter(name='sheet_name', short='s', help='name of worksheet')
	@args.parameter(name='parent_id', short='p', help='parent LOV group')
	@args.parameter(name='id_col', short='i', help='column name for LOV value ID')
	@args.parameter(name='name_col', short='n', help='column name for LOV value name')
	@args.parameter(name='validation', short='V', default='TEXT', help='attribute validation type')
	@args.parameter(name='length', short='L', type=int, default=256, help='attribute validation length')
	@args.parameter(name='output', short='o', default='step.xml', help='step xml format output file')
	def make_lovs(self, id, name, input, sheet_name=None, parent_id=None, id_col=None, name_col=None, validation=None, length=None, output=None):
		'''
		process an excel file to create a STEP LOV xml import

		| ID | Name |
		|----+------|
		| a  | aye  |
		| b  | bee  |
		|----+------|

		<group_id=sheet_name/>

		'''

		workbook = load_workbook(input)

		if sheet_name:
			sheet = workbook[sheet_name]
		else:
			sheet = workbook.active

		lov = self.make_lov(id, name, validation, length, parent_id)

		rows = list(sheet.rows)

		columns = list(rows[0])
		id_col = id_col or columns[0]
		name_col = name_col or columns[1]

		for r in list(range(len(rows)))[1:]:
			row = rows[r]
			if len(row) > 1:
				value_id = row[columns.index(id_col)].value
				value_name = row[columns.index(name_col)].value
			else:
				value_name = row[columns.index(id_col)].value

			lov.Value.append(ValueType(value_name, ID=value_id))

		self.save(output)


	#_____________________________________________________________
	@args.operation
	@args.parameter(name='input', help='excel file with attribute definitions')
	@args.parameter(name='output', short='o', default='step.xml', help='step xml format output file')
	def process(self, input, output=None):
		'''
		process an excel file to create a STEP import

		| Class  | CSV        | Name | Type    | Length |
		|--------+------------+------+---------+--------|
		| Group1 | group1.csv | id   | varchar | 40     |
		| Group1 | group1.csv | name | varchar | 256    |

		'''

		self.ag = AttributeGroupType(
			ID = self.prefix,
			ShowInWorkbench = 'true',
			ManuallySorted = 'false',
			Name = [
				NameType(self.prefix)
			]
		)
		self.doc.AttributeGroupList.append(self.ag)

		root_id = '%s_Root'%self.prefix
		root_name = '%s Root'%self.prefix

		self.tipe = self.make_type(root_id, root_name, self.parent_type_id)
		self.prod = self.make_prod(root_id, root_name, self.tipe.ID, self.parent_product_id)

		columns = {
			'' : 'text',
			'bigint' : 'integer',
			'date' : 'isodate',
			'datetime' : 'isodatetime',
			'flag': 'text',
			'float' : 'fraction',
			'image url': 'text',
			'int' : 'integer',
			'integer' : 'integer',
			'long' : 'integer',
			'lookup table': 'text',
			'money' : 'number',
			'ntext'  : 'text',
			'number' : 'number',
			'numeric' : 'number',
			'nvarchar' : 'text',
			'relationship' : 'text',
			'string enumeration' : 'text',
			'string' : 'text',
			'thumbnail image url': 'text',
			'unicode' : 'text',
			'varchar' : 'text',
			'varchar2' : 'text',
		}

		workbook = open_workbook(filename=input)

		sheet = workbook.sheets()[0]

		Dictionary = namedtuple(
			'Dictionary',
			list(map(
				lambda x: sheet.cell(0, x).value,
				list(range(sheet.ncols))
			))
		)

		Source = None
		CSV = None

		for r in list(range(sheet.nrows))[1:]:
			cols = list(map(
				lambda x: sheet.cell(r, x).value,
				list(range(sheet.ncols))
			))
			row = Dictionary(*cols)

			if row.Source != Source:
				Source = row.Source
				print('%s'%Source)

				root_path = 'Root_%s'%(row.Source)
				root_id  = '%s_%s'%(self.prefix, self.ids.get(root_path))
				root_name = root_path.replace('_',' ')
				root     = self.make_type(root_id, root_name, self.tipe.ID)
				roots     = self.make_prod(root.ID, '%s'%Source, root.ID, self.prod.ID)

			if row.CSV != CSV:
				CSV = row.CSV

				types_path = 'Node_%s_%s'%(row.Source, row.CSV)
				types_id   = '%s_%s'%(self.prefix, self.ids.get(types_path))
				types_name = types_path.replace('_',' ')
				tipes      = self.make_type(types_id, types_name, root.ID)
				print('\t%s'%types_path)

				type_path = 'Leaf_%s_%s'%(row.Source, row.CSV)
				type_id   = '%s_%s'%(self.prefix, self.ids.get(type_path))
				type_name = type_path.replace('_',' ')
				tipe      = self.make_type(type_id, type_name, tipes.ID)
				print('\t\t%s'%type_path)

				prods = self.make_prod(tipes.ID, '%s*'%row.CSV, tipes.ID, roots.ID)
				prod  = self.make_prod(tipe.ID, '%s'%row.CSV, tipe.ID, prods.ID)

			attr_path = '%s_%s_%s'%(row.Source, row.CSV, row.Name)
			attr_id   = '%s_%s'%(self.prefix, self.ids.get(attr_path))
			attr_name = row.Name
			attr_type = columns[row.Type.lower()]
			length = str(row.Length).split('.')[0]
			if len(length) > 0:
				attr_len  = length
			else:
				attr_len = None

			print('\t\t\t%s'%attr_path)

			attr = self.make_attr(attr_id, attr_name, attr_type, attr_len, self.ags[tipe.ID].ID, tipe.ID)
			prod.AttributeLink.append(
				AttributeLinkType(
					AttributeID = attr.ID
				)
			)

		for attribute in self.attrs.values():
			self.prod.AttributeLink.append(
				AttributeLinkType(
					AttributeID=attribute.ID
				)
			)

		self.save(output)


